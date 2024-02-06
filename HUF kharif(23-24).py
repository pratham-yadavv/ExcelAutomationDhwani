import openpyxl as xl
import pandas as pd
import numpy as np
from datetime import datetime


_3A_path = r'C:/Pratham/python/ExcelAutomation/_3A_kharif.xlsx'
_3B_path = r'C:/Pratham/python/ExcelAutomation/_3B_kharif.xlsx'
_3C_path = r'C:/Pratham/python/ExcelAutomation/_3C_kharif.xlsx'
_4A_path = r'C:/Pratham/python/ExcelAutomation/_4A_kharif.xlsx'
_4B_path = r'C:/Pratham/python/ExcelAutomation/_4B_kharif.xlsx'
#Taking dates from user
start_week_date = input('Enter week start date in DD-MM-YYY format : ')
end_week_date = input('Enter week end date in DD-MM-YYYY format : ')
start_season_date = input('Enter season start date in DD-MM-YYYY format : ')

DMS = r'C:/Pratham/python/ExcelAutomation/151123_Kharif 23-24 Demo Plot & DMS user report & Program dashboard (1).xlsx'
sheet ='2.DMS-UserSummary'
sheet2 ='3.Program Dashboard'
sheet3 ='1.DemoPlot-Summary'

def fill_excel_after_string(string, values_list, excel_file,sheet_name):
    # Load the Excel workbook
    wb = xl.load_workbook(excel_file)
    # Select the specific sheet (sheet2)
    sheet = wb[sheet_name]

    # Search for the cell containing the given string
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == string:
                # Get the row and column indexes of the cell containing the string
                row_index = cell.row
                col_index = cell.column

                # Write values from the list just after the cell containing the string
                for i, value in enumerate(values_list):
                    sheet.cell(row=row_index, column=col_index+i+1, value=value)

                # Save the changes to the Excel file
                wb.save(excel_file)
                return

    print(f"String '{string}' not found in sheet of the Excel file.")



def convert_date(date_obj):
    date_str=str(date_obj)
    try:
        # Try parsing the date using the first format ("%d-%m-%Y %H:%M:%S")
        date_a = datetime.strptime(date_str, "%d-%m-%Y %H:%M:%S")
    except ValueError:
        try:
            # If parsing fails, try the second format ("%d/%m/%Y %H:%M")
            date_a = datetime.strptime(date_str, "%Y-%d-%m %H:%M:%S")
        except ValueError:
            return None
    return date_a.strftime('%d-%m-%Y')

def splitter(count_obj) :
    count = len(count_obj.split("||"))
    return count







list1 = [_3A_path,_3B_path,_3C_path,_4A_path,_4B_path]
list2 = ['CIPT','Srijan','Viksat','Pani','Pradan','SSP','SIED']
test_data = ['CIPT Test','SRIJAN Test','VIKSAT Test','PANI Test','PRADAN Test','SSP Test','SIED Test']

list_season_3A_partner = []
list_week_3A_partner = []
list_season_3B_partner = []
list_week_3B_partner = []
list_season_3C_partner = []
list_week_3C_partner = []



for i in list1:
    df = pd.read_excel(i) #Load Data
    df['Server Synced At']=df['Server Synced At'].astype(str)
    df['Server Synced At'] = pd.to_datetime(df['Server Synced At'].apply(convert_date),format='%d-%m-%Y')
    df['Server Synced At'] = df['Server Synced At'].dt.to_pydatetime()
    
    #converting_dateOBj_to_Datetime
    start_date_week = pd.to_datetime(start_week_date,format='%d-%m-%Y')
    end_date_week = pd.to_datetime(end_week_date,format='%d-%m-%Y')
    start_date_season = pd.to_datetime(start_season_date,format='%d-%m-%Y')

    #WEEK_VALUE
    week_df = df[(df['Server Synced At']>=start_date_week)&(df['Server Synced At']<=end_date_week)]
    #SEASON_VALUE
    season_df = df[(df['Server Synced At']>=start_date_season)&(df['Server Synced At']<=end_date_week)]
    # season_value = len(season_df)
    
    if i == _3A_path:
        _3A_list_week = []
        _3A_list_season=[]

        season_df = season_df.drop(season_df[season_df['Surveyor Name'].isin(test_data)].index)

        for x in list2:
            count_df_3A_week = week_df[week_df['partner']== x]
            df_3aW = count_df_3A_week[['Surveyor Id']]
            list_week_3A_partner.append(df_3aW)

            count_df_3A_season = season_df[season_df['partner']== x]
            df_3aS = count_df_3A_season[['Surveyor Id']]
            list_season_3A_partner.append(df_3aS)

        for x in list2:
            count_PARTNER_week = week_df[week_df['partner']== x].shape[0]
            _3A_list_week.append(count_PARTNER_week)

        for x in list2:
            count_PARTNER_season = season_df[season_df['partner']== x].shape[0]
            _3A_list_season.append(count_PARTNER_season)
        
        #filling values in sheet
        fill_excel_after_string("No. of 3A forms added in the reporting week",_3A_list_week,DMS,sheet)
        fill_excel_after_string("No. of 3A forms added in current season (19th June 2023)",_3A_list_season,DMS,sheet)
        fill_excel_after_string("No. of Programme farmers added in current season",_3A_list_season,DMS,sheet2)
        

        

        print("3A run successfully")
        # +++++++++++++++++++++++++++++++++++++++++++++++++++++
    elif i == _3B_path:
        _3B_list_week =[]
        _3B_list_season=[]
        _3B_list_week_unique =[]
        _3B_list_season_unique=[]
        _3B_list_season_crop =[]
        _3B_partner_sum = []
        
        season_df = season_df.drop(season_df[season_df['Surveyor Name'].isin(test_data)].index)

        for x in list2:
            count_df_3B_week = week_df[week_df['partner']== x]
            df_3bW = count_df_3B_week[['Surveyor Id']]
            list_week_3B_partner.append(df_3bW)

            count_df_3B_season = season_df[season_df['partner']== x]
            df_3bS = count_df_3B_season[['Surveyor Id']]
            list_season_3B_partner.append(df_3bS)


        for x in list2:
            count_PARTNER_week = week_df[week_df['partner']== x].shape[0]
            _3B_list_week.append(count_PARTNER_week)


        for x in list2:
            count_PARTNER_season = season_df[season_df['partner']== x].shape[0]
            _3B_list_season.append(count_PARTNER_season)

        for x in list2:
            PARTNER_week = week_df[week_df['partner']== x]
            _unique_week_value_PARTNER = PARTNER_week['Generate Unique ID for Farmer'].nunique()
            _3B_list_week_unique.append(_unique_week_value_PARTNER)

        for x in list2:
            PARTNER_season = season_df[season_df['partner']== x]
            _unique_season_value_PARTNER = PARTNER_season['Generate Unique ID for Farmer'].nunique()
            _3B_list_season_unique.append(_unique_season_value_PARTNER)
            total_sum = PARTNER_season['In how much area are you adopting new and improved farming practices for this Crop?'].sum()
            
            if x == 'CIPT':
                total_sum = total_sum*0.4
                _3B_partner_sum.append(total_sum)
            elif x == 'Srijan':
                total_sum = total_sum*0.16
                _3B_partner_sum.append(total_sum)
            elif x == 'Viksat':
                total_sum = total_sum*0.01
                _3B_partner_sum.append(total_sum)
            elif x=='Pani':
                total_sum = total_sum*0.08
                _3B_partner_sum.append(total_sum)
            elif x == 'Pradan':
                total_sum = total_sum*0.13
                _3B_partner_sum.append(total_sum)
            elif x == 'SSP':
                total_sum = total_sum*0.4
                _3B_partner_sum.append(total_sum)
            elif x == 'SIED':
                total_sum = total_sum*0.4
                _3B_partner_sum.append(total_sum)



        for x in list2:
            count_PARTNER_season = season_df[(season_df['partner']== x)&(season_df['Is this Crop Card - Programme Plot? ']=='Yes')].shape[0]
            _3B_list_season_crop.append(count_PARTNER_season)

        
        #data filling for unique week values
        fill_excel_after_string("No. of 3B Programme Farmers added in the reporting week",_3B_list_week_unique,DMS,sheet)

        #data filling for unique season values
        fill_excel_after_string("No. of 3B Programme Farmers added in current season",_3B_list_season_unique,DMS,sheet)

        #data filling for week values
        fill_excel_after_string("No. of 3B Programme Plots added in the reporting week",_3B_list_week,DMS,sheet)
        
        #data filling for season value
        fill_excel_after_string("No. of 3B Programme Plots added in the current season",_3B_list_season,DMS,sheet)

        #data filling for season crop - Yes
        fill_excel_after_string("No. of 3B forms identified as Programme Crop Card in current season",_3B_list_season_crop,DMS,sheet)
        
        #PROGRAM DASHBOARD
        fill_excel_after_string("# No. of Adoption Farmers",_3B_list_season_unique,DMS,sheet2)
        fill_excel_after_string("# No. of Programme Plots",_3B_list_season,DMS,sheet2)
        fill_excel_after_string("Area adopted under improved practices (Programme Coverage)",_3B_partner_sum,DMS,sheet2)

        print("3B run successfully")

    elif i == _3C_path:
        _3C_week = []
        _3C_PT_week=[]
        _3C_CT_week =[]
        _3C_season =[]
        _3C_PT_season = []
        _3C_CT_season = []
        _3C_unique_PT = []
        _3C_unique_CT = []

        season_df = season_df.drop(season_df[season_df['Surveyor Name'].isin(test_data)].index)

        for x in list2:
            count_df_3C_week = week_df[week_df['partner']== x]
            df_3cW = count_df_3C_week[['Surveyor Id']]
            list_week_3C_partner.append(df_3cW)

            count_df_3C_season = season_df[season_df['partner']== x]
            df_3cS = count_df_3C_season[['Surveyor Id']]
            list_season_3C_partner.append(df_3cS)

        for x in list2:
            count_PARTNER_week = week_df[week_df['partner']== x].shape[0]
            _3C_week.append(count_PARTNER_week)

        for x in list2:
            count_PARTNER_week_PT = week_df[(week_df['partner']== x)&(week_df['Select Plot Category   ']=='Programme Plot')].shape[0]
            _3C_PT_week.append(count_PARTNER_week_PT)

        for x in list2:
            count_PARTNER_week_CT = week_df[(week_df['partner']== x)&(week_df['Select Plot Category   ']=='Control Plot')].shape[0]
            _3C_CT_week.append(count_PARTNER_week_CT)

        for x in list2:
            count_PARTNER_season = season_df[season_df['partner']== x].shape[0]
            _3C_season.append(count_PARTNER_season)

        for x in list2:
            count_PARTNER_season_PT = season_df[(season_df['partner']== x)&(season_df['Select Plot Category   ']=='Programme Plot')].shape[0]
            _3C_PT_season.append(count_PARTNER_season_PT)

            #3.PROGRAM DASHBOARD
            df_PT_ = season_df[(season_df['partner']== x)&(season_df['Select Plot Category   ']=='Programme Plot')]
            count_df_PT = df_PT_['Programme Farmer Unique ID'].nunique()
            _3C_unique_PT.append(count_df_PT)
        for x in list2:
            count_PARTNER_season_CT = season_df[(season_df['partner']== x)&(season_df['Select Plot Category   ']=='Control Plot')].shape[0]
            _3C_CT_season.append(count_PARTNER_season_CT)

            #3.PROGRAM DASHBOARD
            df_CT_ = season_df[(season_df['partner']== x)&(season_df['Select Plot Category   ']=='Control Plot')]
            count_df_CT = df_CT_['Control Farmer Unique ID'].nunique()
            _3C_unique_CT.append(count_df_CT)

        #data filling in week  
        fill_excel_after_string("No. of 3C Crop Cards added in reporting week",_3C_week,DMS,sheet)
        
        #data filling in PT week
        fill_excel_after_string("No. of 3C Programme Crop Cards added in reporting week",_3C_PT_week,DMS,sheet)

        #data filing in CT week
        fill_excel_after_string("No. of 3C Control Crop Cards added in reporting week",_3C_CT_week,DMS,sheet)

        #data filling in season 
        fill_excel_after_string("No. of 3C Crop Cards added in current season",_3C_season,DMS,sheet)

        #data filling in PT season
        fill_excel_after_string("No. of 3C Programme Crop Cards added in current season",_3C_PT_season,DMS,sheet)

        #data filing in CT Season
        fill_excel_after_string("No. of 3C Control Crop Cards added in current season",_3C_CT_season,DMS,sheet)

        #PROGRAM DASHBOARD
        fill_excel_after_string("# No. of Program Plot Crop Card Plots",_3C_PT_season,DMS,sheet2)         
        fill_excel_after_string("# No. of Control Plots Crop Card Farmer",_3C_CT_season,DMS,sheet2)
        fill_excel_after_string("# No. of Program Plot Crop card Farmers",_3C_unique_PT,DMS,sheet2)
        fill_excel_after_string("# No. of Control Plot Crop Card Farmers",_3C_unique_CT,DMS,sheet2)
        print("3C run successfully")

    elif i == _4A_path:
        _4A_season =[]
        _4A_unique_season = []
        list_week_4A_partner = []
        list_4A_sum = []

        season_df = season_df.drop(season_df[season_df['Surveyor Name'].isin(test_data)].index)



        for x in list2:
            count_PARTNER_season_4A = season_df[season_df['partner']== x].shape[0]
            _4A_season.append(count_PARTNER_season_4A)
        
        for x in list2:
            PARTNER_season_4A_unique = season_df[season_df['partner']== x]
            _unique_season_value_PARTNER = PARTNER_season_4A_unique['Surveyor Id'].nunique()
            _4A_unique_season.append(_unique_season_value_PARTNER)

        for x in list2:
            PARTNER_season_4A_sum = season_df[season_df['partner']== x]
            value_4A = 0
            

            for i in range(35,48,2):
                
                df_sum_4A=PARTNER_season_4A_sum.iloc[:,i].to_frame()
                if i == 35:
                    count_1 = df_sum_4A.dropna(subset=['If yes select the planned activities'])
                    count_1 = count_1['If yes select the planned activities'].apply(splitter).sum()
                    value_4A = value_4A + count_1
                elif i == 37:
                    count_2 = df_sum_4A.dropna(subset=['If yes select the planned activities.1'])
                    count_2 = count_2['If yes select the planned activities.1'].apply(splitter).sum()
                    value_4A = value_4A + count_2
            
                elif i == 39:
                    count_3 = df_sum_4A.dropna(subset=['If yes select the planned activities.2'])
                    count_3 = count_3['If yes select the planned activities.2'].apply(splitter).sum()
                    value_4A = value_4A + count_3
                
                elif i == 41:
                    count_4 = df_sum_4A.dropna(subset=['If yes select the planned activities.3'])
                    count_4 = count_4['If yes select the planned activities.3'].apply(splitter).sum()
                    value_4A = value_4A + count_4
                    
                    
                elif i == 43:
                    count_5 = df_sum_4A.dropna(subset=['If yes select the planned activities.4'])
                    count_5 = count_5['If yes select the planned activities.4'].apply(splitter).sum()
                    value_4A = value_4A + count_5
            
                elif i == 45:
                    count_6 = df_sum_4A.dropna(subset=['If yes select the planned activities.5'])
                    count_6 = count_6['If yes select the planned activities.5'].apply(splitter).sum()
                    value_4A = value_4A + count_6
                    
                elif i == 47:
                    count_7 = df_sum_4A.dropna(subset=['If yes select the planned activities.6'])
                    count_7 = count_7['If yes select the planned activities.6'].apply(splitter).sum()
                    value_4A = value_4A + count_7
            list_4A_sum.append(value_4A)
        

    
        for x in list2:
            count_df_4A_week = week_df[week_df['partner']== x]
            df_3cW = count_df_4A_week[['Surveyor Id']]
            list_week_4A_partner.append(df_3cW)

        

        fill_excel_after_string("No. of CRPs who have registered and planned Demo Plots in current season",_4A_unique_season,DMS,sheet3)
        fill_excel_after_string("No. of Demo Plots Registered and planned  in current season",_4A_season,DMS,sheet3)
        fill_excel_after_string("No. of Demo Activities Planned in Demo plots in current season",list_4A_sum,DMS,sheet3)
        print("4A run successfully")
        

    elif i== _4B_path:
        _4B_week =[]
        _4B_unique_week = []
        _4B_season =[]
        _4B_unique_season = []
        _4B_farmerSUM_week = []
        _4B_farmerSUM_season = []
        list_4B_sum_week =[]
        list_4B_sum_season =[]
        list_week_4B_partner=[]
        
        season_df = season_df.drop(season_df[season_df['Surveyor Name'].isin(test_data)].index)


        for x in list2:
            count_PARTNER_season_4B = season_df[season_df['partner']== x].shape[0]
            _4B_season.append(count_PARTNER_season_4B)
        
        for x in list2:
            PARTNER_season_4B_unique = season_df[season_df['partner']== x]
            _unique_season_value_PARTNER = PARTNER_season_4B_unique['Surveyor Id'].nunique()
            _4B_unique_season.append(_unique_season_value_PARTNER)

        for x in list2:
            count_PARTNER_week_4B = week_df[week_df['partner']== x].shape[0]
            _4B_week.append(count_PARTNER_week_4B)
        
        for x in list2:
            PARTNER_week_4B_unique = week_df[week_df['partner']== x]
            _unique_week_value_PARTNER = PARTNER_week_4B_unique['Surveyor Id'].nunique()
            _4B_unique_week.append(_unique_week_value_PARTNER)

        for x in list2:
            sum_farmer_week = week_df[week_df['partner']== x]
            sum_farmer_week_value = sum_farmer_week['Number of farmers who attended the Demo'].sum()
            _4B_farmerSUM_week.append(sum_farmer_week_value)

        for x in list2:
            sum_farmer_season = season_df[season_df['partner']== x]
            sum_farmer_season_value = sum_farmer_season['Number of farmers who attended the Demo'].sum()
            _4B_farmerSUM_season.append(sum_farmer_season_value)

        
        for x in list2:
            PARTNER_week_4B_sum = week_df[week_df['partner']== x]
            value_4A = 0
            

            for i in range(45,52):
                
                df_sum_4B=PARTNER_week_4B_sum.iloc[:,i].to_frame()
                if i == 45:
                    count_1 = df_sum_4B.dropna(subset=['Land Preparation: Select demonstration activities'])
                    count_1 = count_1['Land Preparation: Select demonstration activities'].apply(splitter).sum()
                    value_4A = value_4A + count_1
                elif i == 46:
                    count_2 = df_sum_4B.dropna(subset=['Seed Treatment and Sowing: Select demonstration activities'])
                    count_2 = count_2['Seed Treatment and Sowing: Select demonstration activities'].apply(splitter).sum()
                    value_4A = value_4A + count_2
            
                elif i == 47:
                    count_3 = df_sum_4B.dropna(subset=['Soil Health and Nutrition: Select demonstration activities'])
                    count_3 = count_3['Soil Health and Nutrition: Select demonstration activities'].apply(splitter).sum()
                    value_4A = value_4A + count_3
                
                elif i == 48:
                    count_4 = df_sum_4B.dropna(subset=['Plant Growth: Select demonstration activities'])
                    count_4 = count_4['Plant Growth: Select demonstration activities'].apply(splitter).sum()
                    value_4A = value_4A + count_4
                    
                    
                elif i == 49:
                    count_5 = df_sum_4B.dropna(subset=['Pest and Weed Management: Select demonstration activities'])
                    count_5 = count_5['Pest and Weed Management: Select demonstration activities'].apply(splitter).sum()
                    value_4A = value_4A + count_5
            
                elif i == 50:
                    count_6 = df_sum_4B.dropna(subset=['Irrigation: Select demonstration activities'])
                    count_6 = count_6['Irrigation: Select demonstration activities'].apply(splitter).sum()
                    value_4A = value_4A + count_6
                    
                elif i == 51:
                    count_7 = df_sum_4B.dropna(subset=['Post Harvesting: Select demonstration activities'])
                    count_7 = count_7['Post Harvesting: Select demonstration activities'].apply(splitter).sum()
                    value_4A = value_4A + count_7
            list_4B_sum_week.append(value_4A)



        for x in list2:
            PARTNER_week_4B_sum = season_df[season_df['partner']== x]
            value_4A = 0
            

            for i in range(45,52):
                
                df_sum_4B=PARTNER_week_4B_sum.iloc[:,i].to_frame()
                if i == 45:
                    count_8 = df_sum_4B.dropna(subset=['Land Preparation: Select demonstration activities'])
                    count_8 = count_8['Land Preparation: Select demonstration activities'].apply(splitter).sum()
                    value_4A = value_4A + count_8
                elif i == 46:
                    count_9 = df_sum_4B.dropna(subset=['Seed Treatment and Sowing: Select demonstration activities'])
                    count_9 = count_9['Seed Treatment and Sowing: Select demonstration activities'].apply(splitter).sum()
                    value_4A = value_4A + count_9
            
                elif i == 47:
                    count_10 = df_sum_4B.dropna(subset=['Soil Health and Nutrition: Select demonstration activities'])
                    count_10 = count_10['Soil Health and Nutrition: Select demonstration activities'].apply(splitter).sum()
                    value_4A = value_4A + count_10
                
                elif i == 48:
                    count_11 = df_sum_4B.dropna(subset=['Plant Growth: Select demonstration activities'])
                    count_11 = count_11['Plant Growth: Select demonstration activities'].apply(splitter).sum()
                    value_4A = value_4A + count_11
                    
                    
                elif i == 49:
                    count_12 = df_sum_4B.dropna(subset=['Pest and Weed Management: Select demonstration activities'])
                    count_12 = count_12['Pest and Weed Management: Select demonstration activities'].apply(splitter).sum()
                    value_4A = value_4A + count_12
            
                elif i == 50:
                    count_13 = df_sum_4B.dropna(subset=['Irrigation: Select demonstration activities'])
                    count_13 = count_13['Irrigation: Select demonstration activities'].apply(splitter).sum()
                    value_4A = value_4A + count_13
                    
                elif i == 51:
                    count_14 = df_sum_4B.dropna(subset=['Post Harvesting: Select demonstration activities'])
                    count_14 = count_14['Post Harvesting: Select demonstration activities'].apply(splitter).sum()
                    value_4A = value_4A + count_13
            list_4B_sum_season.append(value_4A)


        
        for x in list2:
            count_df_4B_week = week_df[week_df['partner']== x]  
            df_3cW = count_df_4B_week[['Surveyor Id']]
            list_week_4B_partner.append(df_3cW)

        
        fill_excel_after_string("No. of CRPs who have conducted Demos in reporting week",_4B_unique_week,DMS,sheet3)
        fill_excel_after_string("No. of Demos conducted in reporting week",_4B_week,DMS,sheet3)
        fill_excel_after_string("No. of Demo Activities conducted  in reporting week",list_4B_sum_week,DMS,sheet3)
        fill_excel_after_string("No. of farmers who attended demos in reporting week",_4B_farmerSUM_week,DMS,sheet3)
        fill_excel_after_string("No. of CRPs who have conducted Demos in Current Season",_4B_unique_season,DMS,sheet3)
        fill_excel_after_string("No. of Demos conducted in current season",_4B_season,DMS,sheet3)
        fill_excel_after_string("No. of Demo Activities conducted in current season",list_4B_sum_season,DMS,sheet3)
        fill_excel_after_string("No. of farmers who attended demos in current season",_4B_farmerSUM_season,DMS,sheet3)

        print("4B run successfully")

unique_value_week_3A3B3C = []
unique_value_season_3A3B3C = []
unique_value_week_4A4B = [] 
for i in range(0,7):
    df_week_C3A3B3C = pd.concat([list_week_3A_partner[i],list_week_3B_partner[i],list_week_3C_partner[i]])
    a = df_week_C3A3B3C['Surveyor Id'].nunique()
    unique_value_week_3A3B3C.append(a)

for i in range(0,7):
    df_season_C = pd.concat([list_season_3A_partner[i],list_season_3B_partner[i],list_season_3C_partner[i]])
    a = df_season_C['Surveyor Id'].nunique()
    unique_value_season_3A3B3C.append(a)


for i in range(0,7):
    df_week_C4A4B = pd.concat([list_week_4A_partner[i],list_week_4B_partner[i]])
    a = df_week_C4A4B['Surveyor Id'].nunique()
    unique_value_week_4A4B.append(a)


fill_excel_after_string("# of Active CRPs/Users IDs in reporting week ",unique_value_week_3A3B3C,DMS,sheet)
fill_excel_after_string("# of Active CRPs/Users IDs in Current Season",unique_value_season_3A3B3C,DMS,sheet)
fill_excel_after_string("# of Active CRPs/Users IDs in reporting week ",unique_value_week_4A4B,DMS,sheet3)

print("Hurray ! Program run Successfully")